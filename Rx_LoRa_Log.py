from openpyxl import Workbook
import serial
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
import argparse


def main():
    parser = argparse.ArgumentParser(description='Process name and number arguments.')
    parser.add_argument('NAME', type=str, help='Provide a name.')
    parser.add_argument('Number', type=int, help='Provide a number.')

    args = parser.parse_args()

    location = args.NAME
    sf = args.Number
    file_name = f"{location}_{sf}.xlsx"

    print(file_name,flush=True)

    workbook = Workbook()
    sheet = workbook.active

    sheet.merge_cells('A1:H1')
    cell = sheet["A1"]
    cell.font = Font(size=10, bold=True, italic=True, underline='single', color='008000')
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.value = "Location: " + str(location)

    sheet.merge_cells('A2:H2')
    cell = sheet["A2"]
    cell.font = Font(size=10, bold=True, italic=True, underline='single', color='008000')
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.value = "SF: " + str(sf)

    sheet.merge_cells('A3:H3')
    cell = sheet["A3"]
    cell.font = Font(size=10, bold=True, italic=True, underline='single', color='008000')
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.value = " "
    row_names = ["A4", "B4", "C4", "D4", "E4", "F4", "G4", "H4"]
    titles = ["Received Data", "RSSI", "SNR", "CRC Error", "Byte 1", "Byte 2", "Byte 3", "Byte 4"]
    for i in range(8):
        cell = sheet[row_names[i]]
        cell.font = Font(size=10, bold=True, italic=True, underline='single', color='008000')
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.value = titles[i]

    x = 0
    ser = serial.Serial('/dev/ttyACM0', 9600, timeout=10)
    cycle_ongoing = True
    while cycle_ongoing == True:
        rx_data = ser.readline().decode('utf-8').rstrip()
        print(rx_data,flush=True)
        if rx_data.strip() == "Set_SF:":
            send_char = str(sf)
            if sf == 10:
                send_char = str('a')
            elif sf == 11:
                send_char = str('b')
            elif sf == 12:
                send_char = str('c')
            byte_data = send_char.encode('utf-8')
#            ser.write(byte_data)
            print("Data sent from RPi",flush=True)
        x = x + 1
        if "dBm" in rx_data:
            log_data = rx_data.split(" ")
            rssi_tmp = log_data[0].replace('rssi:', '')
            rssi = rssi_tmp.replace('dBm', '')
            snr_tmp = log_data[2].replace('snr:', '')
            snr = snr_tmp.replace('dB', '')
            crc_err = log_data[4].replace('crc_error:', '')
            data = (rx_data, rssi, snr, crc_err, log_data[6], log_data[7], log_data[8], log_data[9])
            sheet.append(data)
            workbook.save(filename=file_name)
        if "CycleCompleted" in rx_data:
            print("One Cycle Finished",flush=True)
            cycle_ongoing = False

    workbook.save


if __name__ == '__main__':
    main()
